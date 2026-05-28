#!/usr/bin/env python3
"""
Concept2 Logbook to Excel - Per Person Sheets

Creates a separate sheet for each person with their L/R side workouts
displayed side by side, with highlighted peak power.
"""

import os
import sys
from pathlib import Path
from collections import defaultdict

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from concept2_client import Concept2Client, Workout


# Styling
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
PEAK_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LEFT_HEADER = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
RIGHT_HEADER = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

# Column layout: LEFT in cols 1-2, gap in col 3, RIGHT in cols 4-5
LEFT_COL = 1
RIGHT_COL = 4


def write_person_sheet(ws, person: str, workouts: list[Workout]):
    """Write a person's workouts side by side (LEFT | RIGHT)."""
    left_workouts = [w for w in workouts if w.side == 'L']
    right_workouts = [w for w in workouts if w.side == 'R']

    row = 1

    # Headers - only show if data exists for that side
    if left_workouts:
        cell = ws.cell(row=row, column=LEFT_COL, value="LEFT SIDE")
        cell.font = Font(bold=True, size=14)
        cell.fill = LEFT_HEADER

    if right_workouts:
        cell = ws.cell(row=row, column=RIGHT_COL, value="RIGHT SIDE")
        cell.font = Font(bold=True, size=14)
        cell.fill = RIGHT_HEADER

    row += 2

    # Write workouts side by side
    max_workouts = max(len(left_workouts), len(right_workouts), 1)

    for i in range(max_workouts):
        left_w = left_workouts[i] if i < len(left_workouts) else None
        right_w = right_workouts[i] if i < len(right_workouts) else None

        if not left_w and not right_w:
            continue

        # Save start row for both columns
        start_row = row

        # Write LEFT workout
        if left_w:
            left_end = write_workout_column(ws, start_row, LEFT_COL, left_w)
        else:
            left_end = start_row

        # Write RIGHT workout at same starting row
        if right_w:
            right_end = write_workout_column(ws, start_row, RIGHT_COL, right_w)
        else:
            right_end = start_row

        # Move to next section (use whichever ended lower)
        row = max(left_end, right_end) + 2

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 3  # Gap
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12


def write_workout_column(ws, start_row: int, col: int, workout: Workout) -> int:
    """Write a single workout in a column. Returns next row."""
    row = start_row

    # Date
    cell = ws.cell(row=row, column=col, value=f"Date: {workout.date}")
    cell.font = HEADER_FONT
    row += 1

    # Avg power
    ws.cell(row=row, column=col, value="Avg Power:")
    ws.cell(row=row, column=col + 1, value=workout.avg_power or "N/A")
    row += 1

    # Peak power
    ws.cell(row=row, column=col, value="Peak Power:")
    peak_cell = ws.cell(row=row, column=col + 1, value=workout.peak_power or "N/A")
    peak_cell.font = Font(bold=True)
    row += 1

    # Stroke header
    ws.cell(row=row, column=col, value="Stroke #").font = HEADER_FONT
    ws.cell(row=row, column=col).fill = HEADER_FILL
    ws.cell(row=row, column=col + 1, value="Power (W)").font = HEADER_FONT
    ws.cell(row=row, column=col + 1).fill = HEADER_FILL
    row += 1

    # Stroke data
    for stroke in workout.strokes:
        ws.cell(row=row, column=col, value=stroke.index)
        power_cell = ws.cell(row=row, column=col + 1, value=stroke.power)

        if stroke.index == workout.peak_index:
            ws.cell(row=row, column=col).fill = PEAK_FILL
            power_cell.fill = PEAK_FILL
            power_cell.font = Font(bold=True)

        row += 1

    return row


def write_empty_column(ws, start_row: int, col: int, pad_to: int) -> int:
    """Write empty placeholder column. Returns next row."""
    row = start_row + 4  # Skip header rows
    row += pad_to
    return row


def main():
    data_dir = Path(__file__).parent.parent / "data"
    env_path = data_dir / ".env"
    if not env_path.exists():
        print("Missing data/.env file.")
        sys.exit(1)

    load_dotenv(env_path)

    c2_client_id = os.getenv("C2_CLIENT_ID")
    c2_client_secret = os.getenv("C2_CLIENT_SECRET")
    c2_redirect_uri = os.getenv("C2_REDIRECT_URI", "http://localhost:8080/callback")

    if not c2_client_id or not c2_client_secret:
        print("Missing Concept2 API credentials in .env")
        sys.exit(1)

    print("\n=== Concept2 Authentication ===")
    c2 = Concept2Client(c2_client_id, c2_client_secret, c2_redirect_uri)
    if not c2.authenticate():
        print("Failed to authenticate with Concept2")
        sys.exit(1)

    print("\n=== Fetching Workouts ===")
    workouts = c2.get_workouts()

    if not workouts:
        print("No workouts found")
        sys.exit(0)

    # Group by person (case-insensitive)
    by_person = defaultdict(list)
    for w in workouts:
        name = w.person.strip().title()
        by_person[name].append(w)

    print(f"\nFound {len(workouts)} workouts for {len(by_person)} people:")
    for person, person_workouts in by_person.items():
        left = sum(1 for w in person_workouts if w.side == 'L')
        right = sum(1 for w in person_workouts if w.side == 'R')
        print(f"  - {person}: {left} L, {right} R")

    print("\n=== Writing to Excel ===")
    wb = Workbook()
    wb.remove(wb.active)

    for person, person_workouts in sorted(by_person.items()):
        sheet_name = person[:31].replace('/', '-').replace('\\', '-')
        ws = wb.create_sheet(title=sheet_name)
        write_person_sheet(ws, person, person_workouts)

    output_path = data_dir / "workouts.xlsx"
    wb.save(output_path)
    print(f"Saved to {output_path}")

    print("\n=== Done! ===")


if __name__ == "__main__":
    main()
