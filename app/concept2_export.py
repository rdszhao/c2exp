#!/usr/bin/env python3
"""
Concept2 Logbook Exporter - Standalone Windows App

Exports workout data from Concept2 Logbook to Excel with per-person sheets,
showing LEFT/RIGHT sides with stroke data and highlighted peak power.
"""

import json
import os
import re
import sys
import threading
import webbrowser
from dataclasses import dataclass, field
from http.server import HTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from typing import Optional
from urllib.parse import urlencode, urlparse, parse_qs

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

# ============================================================================
# CONCEPT2 API CREDENTIALS (Southbreeze Paddle)
# ============================================================================
C2_CLIENT_ID = "bLBbRvrMCEFQcXFZJBonQCkxRlxSziqpyKV9hRGQ"
C2_CLIENT_SECRET = "u7TZpJgvhHyJDVPPvTvKs9UhMPwnlMvAUxmMHnYt"
C2_REDIRECT_URI = "http://localhost:8080/callback"

BASE_URL = "https://log.concept2.com"
AUTH_URL = f"{BASE_URL}/oauth/authorize"
TOKEN_URL = f"{BASE_URL}/oauth/access_token"


def get_app_data_dir() -> Path:
    """Get platform-appropriate app data directory."""
    if sys.platform == "win32":
        base = Path(os.environ.get("APPDATA", Path.home()))
    elif sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    else:
        base = Path.home() / ".config"

    app_dir = base / "Concept2Export"
    app_dir.mkdir(parents=True, exist_ok=True)
    return app_dir


def get_output_dir() -> Path:
    """Get user's Documents folder for output."""
    if sys.platform == "win32":
        docs = Path(os.environ.get("USERPROFILE", Path.home())) / "Documents"
    else:
        docs = Path.home() / "Documents"

    if not docs.exists():
        docs = Path.home()
    return docs


TOKEN_FILE = get_app_data_dir() / "token.json"


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class Stroke:
    index: int
    power: float


@dataclass
class Workout:
    date: str
    person: str
    side: str
    strokes: list[Stroke] = field(default_factory=list)
    avg_power: Optional[float] = None
    peak_power: Optional[float] = None
    peak_index: Optional[int] = None

    @staticmethod
    def parse_notes(notes: str) -> tuple[str, str]:
        if not notes:
            return "Unknown", ""
        notes = notes.strip()
        match = re.match(r'^(.+?)\s+([LR])\b', notes, re.IGNORECASE)
        if match:
            return match.group(1).strip(), match.group(2).upper()
        if notes.endswith(' L') or notes.endswith(' l'):
            return notes[:-2].strip(), 'L'
        if notes.endswith(' R') or notes.endswith(' r'):
            return notes[:-2].strip(), 'R'
        return notes, ""


# ============================================================================
# OAUTH HANDLER
# ============================================================================

class OAuthCallbackHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        query = parse_qs(urlparse(self.path).query)
        if "code" in query:
            self.server.auth_code = query["code"][0]
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(b"""
                <html><body style="font-family: Arial; text-align: center; padding: 50px;">
                <h1 style="color: green;">Authorization Successful!</h1>
                <p>You can close this window and return to the app.</p>
                </body></html>
            """)
        else:
            self.server.auth_code = None
            self.send_response(400)
            self.end_headers()
            self.wfile.write(b"Authorization failed")

    def log_message(self, format, *args):
        pass


# ============================================================================
# CONCEPT2 CLIENT
# ============================================================================

class Concept2Client:
    def __init__(self, status_callback=None):
        self.status_callback = status_callback or print
        self.access_token: Optional[str] = None
        self.refresh_token: Optional[str] = None
        self.user_id: Optional[int] = None
        self._load_token()

    def _status(self, msg: str):
        self.status_callback(msg)

    def _load_token(self):
        if TOKEN_FILE.exists():
            try:
                data = json.loads(TOKEN_FILE.read_text())
                self.access_token = data.get("access_token")
                self.refresh_token = data.get("refresh_token")
                self.user_id = data.get("user_id")
            except (json.JSONDecodeError, KeyError):
                pass

    def _save_token(self):
        TOKEN_FILE.write_text(json.dumps({
            "access_token": self.access_token,
            "refresh_token": self.refresh_token,
            "user_id": self.user_id,
        }))

    def authenticate(self) -> bool:
        if self.access_token and self._test_token():
            self._status("Using saved login credentials...")
            return True

        self._status("Opening browser for Concept2 login...")

        params = {
            "client_id": C2_CLIENT_ID,
            "redirect_uri": C2_REDIRECT_URI,
            "response_type": "code",
        }
        auth_url = f"{AUTH_URL}?{urlencode(params)}"

        parsed = urlparse(C2_REDIRECT_URI)
        port = parsed.port or 8080

        try:
            server = HTTPServer(("localhost", port), OAuthCallbackHandler)
        except OSError:
            self._status(f"Error: Port {port} is in use. Close other apps and retry.")
            return False

        server.auth_code = None
        server.timeout = 120  # 2 minute timeout

        webbrowser.open(auth_url)
        self._status("Waiting for authorization (check your browser)...")

        server.handle_request()

        if not server.auth_code:
            self._status("Authorization failed - no code received")
            return False

        self._status("Exchanging authorization code...")

        response = requests.post(TOKEN_URL, data={
            "grant_type": "authorization_code",
            "client_id": C2_CLIENT_ID,
            "client_secret": C2_CLIENT_SECRET,
            "redirect_uri": C2_REDIRECT_URI,
            "code": server.auth_code,
        })

        if response.status_code != 200:
            self._status(f"Token exchange failed: {response.text}")
            return False

        data = response.json()
        self.access_token = data["access_token"]
        self.refresh_token = data.get("refresh_token")

        if not self._fetch_user_id():
            return False

        self._save_token()
        self._status("Login successful!")
        return True

    def _test_token(self) -> bool:
        if not self.access_token:
            return False
        try:
            resp = self._api_get("/api/users/me")
            if resp is not None:
                return True
            # Token might be expired, try refresh
            if self.refresh_token:
                return self._refresh_access_token()
            return False
        except:
            return False

    def _refresh_access_token(self) -> bool:
        """Use refresh token to get a new access token."""
        if not self.refresh_token:
            return False

        self._status("Refreshing access token...")
        try:
            response = requests.post(TOKEN_URL, data={
                "grant_type": "refresh_token",
                "client_id": C2_CLIENT_ID,
                "client_secret": C2_CLIENT_SECRET,
                "refresh_token": self.refresh_token,
            })

            if response.status_code == 200:
                data = response.json()
                self.access_token = data["access_token"]
                self.refresh_token = data.get("refresh_token", self.refresh_token)
                self._save_token()
                self._status("Token refreshed successfully")
                return True
        except:
            pass

        return False

    def _fetch_user_id(self) -> bool:
        data = self._api_get("/api/users/me")
        if data and "data" in data:
            self.user_id = data["data"]["id"]
            return True
        return False

    def _api_get(self, endpoint: str, params: dict = None) -> Optional[dict]:
        headers = {"Authorization": f"Bearer {self.access_token}"}
        response = requests.get(f"{BASE_URL}{endpoint}", headers=headers, params=params)
        if response.status_code == 200:
            return response.json()
        return None

    def get_workouts(self) -> list[Workout]:
        if not self.user_id:
            if not self._fetch_user_id():
                raise RuntimeError("Could not get user ID")

        workouts = []
        page = 1

        while True:
            self._status(f"Fetching workouts (page {page})...")
            data = self._api_get(
                f"/api/users/{self.user_id}/results",
                params={"per_page": 250, "page": page}
            )

            if not data or "data" not in data:
                break

            for result in data["data"]:
                workout = self._parse_workout(result)
                if workout:
                    workouts.append(workout)

            meta = data.get("meta", {}).get("pagination", {})
            if page >= meta.get("total_pages", 1):
                break
            page += 1

        self._status(f"Found {len(workouts)} workouts")
        return workouts

    def _parse_workout(self, data: dict) -> Optional[Workout]:
        try:
            notes = data.get("comments", "") or ""
            person, side = Workout.parse_notes(notes)

            workout = Workout(
                date=data.get("date", ""),
                person=person,
                side=side,
            )

            self._load_strokes(data["id"], workout)
            return workout
        except (KeyError, TypeError):
            return None

    def _load_strokes(self, result_id: int, workout: Workout):
        if not self.user_id:
            return

        data = self._api_get(f"/api/users/{self.user_id}/results/{result_id}/strokes")
        if not data or "data" not in data:
            return

        strokes_data = data["data"]
        if not strokes_data:
            return

        powers = []
        for i, stroke in enumerate(strokes_data):
            pace_tenths = stroke.get("p", 0)
            if pace_tenths and pace_tenths > 0:
                pace_seconds = pace_tenths / 10.0
                power = 2.80 / ((pace_seconds / 500.0) ** 3)
                power = round(power, 1)
                workout.strokes.append(Stroke(index=i + 1, power=power))
                powers.append((i, power))

        if powers:
            workout.avg_power = round(sum(p for _, p in powers) / len(powers), 1)
            peak_idx, peak_val = max(powers, key=lambda x: x[1])
            workout.peak_power = peak_val
            workout.peak_index = peak_idx + 1


# ============================================================================
# EXCEL EXPORT
# ============================================================================

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
PEAK_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LEFT_HEADER = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
RIGHT_HEADER = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

LEFT_COL = 1
RIGHT_COL = 4


def write_workout_column(ws, start_row: int, col: int, workout: Workout) -> int:
    row = start_row

    cell = ws.cell(row=row, column=col, value=f"Date: {workout.date}")
    cell.font = HEADER_FONT
    row += 1

    ws.cell(row=row, column=col, value="Avg Power:")
    ws.cell(row=row, column=col + 1, value=workout.avg_power or "N/A")
    row += 1

    ws.cell(row=row, column=col, value="Peak Power:")
    peak_cell = ws.cell(row=row, column=col + 1, value=workout.peak_power or "N/A")
    peak_cell.font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=col, value="Stroke #").font = HEADER_FONT
    ws.cell(row=row, column=col).fill = HEADER_FILL
    ws.cell(row=row, column=col + 1, value="Power (W)").font = HEADER_FONT
    ws.cell(row=row, column=col + 1).fill = HEADER_FILL
    row += 1

    for stroke in workout.strokes:
        ws.cell(row=row, column=col, value=stroke.index)
        power_cell = ws.cell(row=row, column=col + 1, value=stroke.power)

        if stroke.index == workout.peak_index:
            ws.cell(row=row, column=col).fill = PEAK_FILL
            power_cell.fill = PEAK_FILL
            power_cell.font = Font(bold=True)

        row += 1

    return row


def write_person_sheet(ws, person: str, workouts: list[Workout]):
    left_workouts = [w for w in workouts if w.side == 'L']
    right_workouts = [w for w in workouts if w.side == 'R']

    row = 1

    if left_workouts:
        cell = ws.cell(row=row, column=LEFT_COL, value="LEFT SIDE")
        cell.font = Font(bold=True, size=14)
        cell.fill = LEFT_HEADER

    if right_workouts:
        cell = ws.cell(row=row, column=RIGHT_COL, value="RIGHT SIDE")
        cell.font = Font(bold=True, size=14)
        cell.fill = RIGHT_HEADER

    row += 2

    max_workouts = max(len(left_workouts), len(right_workouts), 1)

    for i in range(max_workouts):
        left_w = left_workouts[i] if i < len(left_workouts) else None
        right_w = right_workouts[i] if i < len(right_workouts) else None

        if not left_w and not right_w:
            continue

        start_row = row

        if left_w:
            left_end = write_workout_column(ws, start_row, LEFT_COL, left_w)
        else:
            left_end = start_row

        if right_w:
            right_end = write_workout_column(ws, start_row, RIGHT_COL, right_w)
        else:
            right_end = start_row

        row = max(left_end, right_end) + 2

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12


def export_to_excel(workouts: list[Workout], output_path: Path):
    by_person = defaultdict(list)
    for w in workouts:
        name = w.person.strip().title()
        by_person[name].append(w)

    wb = Workbook()
    wb.remove(wb.active)

    for person, person_workouts in sorted(by_person.items()):
        sheet_name = person[:31].replace('/', '-').replace('\\', '-')
        ws = wb.create_sheet(title=sheet_name)
        write_person_sheet(ws, person, person_workouts)

    wb.save(output_path)
    return len(by_person)


# ============================================================================
# GUI APPLICATION
# ============================================================================

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
    HAS_TKINTER = True
except ImportError:
    HAS_TKINTER = False


class WorkoutSelectionDialog:
    """Dialog for selecting which workouts to export."""

    def __init__(self, parent, workouts: list[Workout]):
        self.result = None
        self.workouts = workouts
        self.selected_indices = set(range(len(workouts)))  # All selected by default

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Workouts to Export")
        self.dialog.geometry("700x500")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # Center on parent
        self.dialog.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - 700) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 500) // 2
        self.dialog.geometry(f"700x500+{x}+{y}")

        self._build_ui()

    def _build_ui(self):
        # Header
        header_frame = ttk.Frame(self.dialog)
        header_frame.pack(fill="x", padx=10, pady=10)

        ttk.Label(header_frame, text="Select workouts to export:",
                 font=("Arial", 12, "bold")).pack(side="left")

        self.count_var = tk.StringVar()
        self._update_count()
        ttk.Label(header_frame, textvariable=self.count_var).pack(side="right")

        # Buttons for select all / none
        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(fill="x", padx=10)

        ttk.Button(btn_frame, text="Select All", command=self._select_all).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="Select None", command=self._select_none).pack(side="left", padx=2)

        # Treeview with checkboxes
        tree_frame = ttk.Frame(self.dialog)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side="right", fill="y")

        # Treeview
        columns = ("select", "person", "date", "side", "avg_power", "peak_power")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                  yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.tree.yview)

        # Column headings
        self.tree.heading("select", text="Export")
        self.tree.heading("person", text="Person")
        self.tree.heading("date", text="Date")
        self.tree.heading("side", text="Side")
        self.tree.heading("avg_power", text="Avg Power")
        self.tree.heading("peak_power", text="Peak Power")

        # Column widths
        self.tree.column("select", width=50, anchor="center")
        self.tree.column("person", width=150)
        self.tree.column("date", width=100)
        self.tree.column("side", width=50, anchor="center")
        self.tree.column("avg_power", width=80, anchor="right")
        self.tree.column("peak_power", width=80, anchor="right")

        self.tree.pack(fill="both", expand=True)

        # Populate tree
        for i, w in enumerate(self.workouts):
            check = "✓" if i in self.selected_indices else ""
            self.tree.insert("", "end", iid=str(i), values=(
                check,
                w.person.strip().title(),
                w.date,
                w.side or "-",
                f"{w.avg_power:.1f} W" if w.avg_power else "-",
                f"{w.peak_power:.1f} W" if w.peak_power else "-"
            ))

        # Bind click to toggle selection
        self.tree.bind("<ButtonRelease-1>", self._on_click)

        # Bottom buttons
        bottom_frame = ttk.Frame(self.dialog)
        bottom_frame.pack(fill="x", padx=10, pady=10)

        ttk.Button(bottom_frame, text="Cancel", command=self._cancel).pack(side="right", padx=5)
        ttk.Button(bottom_frame, text="Export Selected", command=self._confirm).pack(side="right", padx=5)

    def _on_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return

        idx = int(item)
        if idx in self.selected_indices:
            self.selected_indices.remove(idx)
            self.tree.set(item, "select", "")
        else:
            self.selected_indices.add(idx)
            self.tree.set(item, "select", "✓")

        self._update_count()

    def _select_all(self):
        self.selected_indices = set(range(len(self.workouts)))
        for i in range(len(self.workouts)):
            self.tree.set(str(i), "select", "✓")
        self._update_count()

    def _select_none(self):
        self.selected_indices = set()
        for i in range(len(self.workouts)):
            self.tree.set(str(i), "select", "")
        self._update_count()

    def _update_count(self):
        total = len(self.workouts)
        selected = len(self.selected_indices)
        self.count_var.set(f"{selected} of {total} selected")

    def _confirm(self):
        self.result = [self.workouts[i] for i in sorted(self.selected_indices)]
        self.dialog.destroy()

    def _cancel(self):
        self.result = None
        self.dialog.destroy()

    def show(self) -> Optional[list[Workout]]:
        self.dialog.wait_window()
        return self.result


class Concept2ExportApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Concept2 Logbook Exporter")
        self.root.geometry("500x400")
        self.root.resizable(False, False)

        # Center window
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() - 500) // 2
        y = (self.root.winfo_screenheight() - 400) // 2
        self.root.geometry(f"500x400+{x}+{y}")

        self._build_ui()
        self.client = None
        self.workouts = []
        self.selected_workouts = []

    def _build_ui(self):
        # Title
        title = ttk.Label(self.root, text="Concept2 Logbook Exporter",
                         font=("Arial", 16, "bold"))
        title.pack(pady=20)

        # Description
        desc = ttk.Label(self.root,
                        text="Export your Concept2 workout data to Excel\n"
                             "with per-person sheets and stroke-by-stroke power data.",
                        justify="center")
        desc.pack(pady=10)

        # Status frame
        status_frame = ttk.LabelFrame(self.root, text="Status", padding=10)
        status_frame.pack(fill="x", padx=20, pady=10)

        self.status_var = tk.StringVar(value="Ready to start")
        self.status_label = ttk.Label(status_frame, textvariable=self.status_var,
                                      wraplength=440)
        self.status_label.pack(fill="x")

        # Progress bar
        self.progress = ttk.Progressbar(status_frame, mode="indeterminate")
        self.progress.pack(fill="x", pady=(10, 0))

        # Buttons frame
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(pady=20)

        self.login_btn = ttk.Button(btn_frame, text="1. Login to Concept2",
                                    command=self._do_login, width=25)
        self.login_btn.pack(pady=5)

        self.fetch_btn = ttk.Button(btn_frame, text="2. Fetch Workouts",
                                    command=self._do_fetch, width=25, state="disabled")
        self.fetch_btn.pack(pady=5)

        self.select_btn = ttk.Button(btn_frame, text="3. Select Workouts",
                                     command=self._do_select, width=25, state="disabled")
        self.select_btn.pack(pady=5)

        self.export_btn = ttk.Button(btn_frame, text="4. Export to Excel",
                                     command=self._do_export, width=25, state="disabled")
        self.export_btn.pack(pady=5)

        # Footer
        footer = ttk.Label(self.root, text="Southbreeze Paddle Erg Tracker",
                          font=("Arial", 9), foreground="gray")
        footer.pack(side="bottom", pady=10)

    def _update_status(self, msg: str):
        self.status_var.set(msg)
        self.root.update_idletasks()

    def _do_login(self):
        self.progress.start()
        self.login_btn.config(state="disabled")

        def login_thread():
            try:
                self.client = Concept2Client(status_callback=self._update_status)
                if self.client.authenticate():
                    self.root.after(0, self._login_success)
                else:
                    self.root.after(0, self._login_failed)
            except Exception as e:
                self.root.after(0, lambda: self._login_error(str(e)))

        threading.Thread(target=login_thread, daemon=True).start()

    def _login_success(self):
        self.progress.stop()
        self._update_status("Logged in successfully!")
        self.fetch_btn.config(state="normal")
        self.login_btn.config(text="1. Login to Concept2 ✓", state="disabled")

    def _login_failed(self):
        self.progress.stop()
        self._update_status("Login failed. Please try again.")
        self.login_btn.config(state="normal")

    def _login_error(self, error: str):
        self.progress.stop()
        self._update_status(f"Error: {error}")
        self.login_btn.config(state="normal")
        messagebox.showerror("Login Error", error)

    def _do_fetch(self):
        self.progress.start()
        self.fetch_btn.config(state="disabled")

        def fetch_thread():
            try:
                self.workouts = self.client.get_workouts()
                self.root.after(0, self._fetch_success)
            except Exception as e:
                self.root.after(0, lambda: self._fetch_error(str(e)))

        threading.Thread(target=fetch_thread, daemon=True).start()

    def _fetch_success(self):
        self.progress.stop()
        count = len(self.workouts)

        # Count unique people
        people = set(w.person.strip().title() for w in self.workouts)

        self._update_status(f"Found {count} workouts from {len(people)} people")
        self.select_btn.config(state="normal")
        self.fetch_btn.config(text="2. Fetch Workouts ✓", state="disabled")

    def _fetch_error(self, error: str):
        self.progress.stop()
        self._update_status(f"Error fetching workouts: {error}")
        self.fetch_btn.config(state="normal")
        messagebox.showerror("Fetch Error", error)

    def _do_select(self):
        dialog = WorkoutSelectionDialog(self.root, self.workouts)
        result = dialog.show()

        if result is None:
            # User cancelled
            return

        if not result:
            messagebox.showwarning("No Selection", "Please select at least one workout to export.")
            return

        self.selected_workouts = result
        people = set(w.person.strip().title() for w in result)
        self._update_status(f"Selected {len(result)} workouts from {len(people)} people")
        self.export_btn.config(state="normal")
        self.select_btn.config(text="3. Select Workouts ✓")

    def _do_export(self):
        default_path = get_output_dir() / "concept2_workouts.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="concept2_workouts.xlsx",
            initialdir=str(get_output_dir()),
            title="Save Workouts As"
        )

        if not filepath:
            return

        self.progress.start()
        self.export_btn.config(state="disabled")

        def export_thread():
            try:
                num_people = export_to_excel(self.selected_workouts, Path(filepath))
                self.root.after(0, lambda: self._export_success(filepath, num_people))
            except Exception as e:
                self.root.after(0, lambda: self._export_error(str(e)))

        threading.Thread(target=export_thread, daemon=True).start()

    def _export_success(self, filepath: str, num_people: int):
        self.progress.stop()
        self._update_status(f"Exported to {filepath}")
        self.export_btn.config(text="4. Export to Excel ✓")

        result = messagebox.askyesno(
            "Export Complete",
            f"Successfully exported {len(self.selected_workouts)} workouts\n"
            f"for {num_people} people to:\n\n{filepath}\n\n"
            "Would you like to open the file?"
        )

        if result:
            if sys.platform == "win32":
                os.startfile(filepath)
            elif sys.platform == "darwin":
                os.system(f'open "{filepath}"')
            else:
                os.system(f'xdg-open "{filepath}"')

    def _export_error(self, error: str):
        self.progress.stop()
        self._update_status(f"Export failed: {error}")
        self.export_btn.config(state="normal")
        messagebox.showerror("Export Error", error)

    def run(self):
        self.root.mainloop()


def console_main():
    """Console-based version for systems without tkinter."""
    print("=" * 50)
    print("Concept2 Logbook Exporter (Console Mode)")
    print("=" * 50)
    print()

    client = Concept2Client(status_callback=print)

    print("Step 1: Authenticating...")
    if not client.authenticate():
        print("Failed to authenticate. Please try again.")
        return

    print()
    print("Step 2: Fetching workouts...")
    workouts = client.get_workouts()

    if not workouts:
        print("No workouts found.")
        return

    people = set(w.person.strip().title() for w in workouts)
    print(f"Found {len(workouts)} workouts from {len(people)} people")
    print()

    output_path = get_output_dir() / "concept2_workouts.xlsx"
    print(f"Step 3: Exporting to {output_path}...")

    num_people = export_to_excel(workouts, output_path)
    print(f"Exported {len(workouts)} workouts for {num_people} people")
    print()
    print(f"File saved to: {output_path}")
    print()
    print("Done!")


def main():
    if HAS_TKINTER:
        app = Concept2ExportApp()
        app.run()
    else:
        print("Note: GUI not available (tkinter not installed)")
        print("Running in console mode...")
        print()
        console_main()


if __name__ == "__main__":
    main()
